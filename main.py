import pandas as pd
import glob
import openpyxl as pyxl
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    df = pd.read_excel(filepath, sheet_name="Sheet 1")
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=False, margin=0)
    pdf.add_page()
    filename = Path(filepath).stem
    invoice_num = filename.split("-")
    pdf.set_font(family="Times", style="B", size=24)
    pdf.cell(w=0, h=12, txt=f"Invoice {invoice_num[0]}", align="L", ln=1)
    pdf.cell(w=0, h=12, txt=f"Date: {invoice_num[1]}", align="L", ln=1)
    pdf.ln(55)

    # Adding Headers
    pdf.set_font(family="Times", style="B", size=8)
    Headings = [item.replace("_", " ").title() for item in df.columns]
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=Headings[0], border=1)
    pdf.cell(w=50, h=8, txt=Headings[1], border=1)
    pdf.cell(w=30, h=8, txt=Headings[2], border=1)
    pdf.cell(w=30, h=8, txt=Headings[3], border=1)
    pdf.cell(w=30, h=8, txt=Headings[4], border=1, ln=1)

    # Adding rows to the table
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=50, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    total_price = sum(df["total_price"])
    pdf.cell(w=140, h=8, txt=" ", border=1)
    pdf.cell(w=30, h=8, txt="£ " + str(total_price), border=1, ln=1)

    # Adding total sum sentence
    pdf.cell(w=30, h=8, txt=f"The total price is £{total_price}", ln=1)

    # Adding company name and logo
    pdf.set_font(family="Times", size=10, style="B")
    pdf.cell(w=25, h=8, txt=f"Gordon Burns LTD")
    pdf.cell(w=6, h=8, txt=" ")
    pdf.image("pythonhow.png", w=9)

    pdf.output(f"PDF_INVOICES/Invoice{invoice_num[0]}.pdf")
